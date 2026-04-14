import os

import pandas as pd
from openpyxl.styles import PatternFill


INPUT_XLSX = "timetable_data.xlsx"
OUTPUT_PREFIX = "final_timetable"


_DEFAULT_TIME_SLOTS = [
    "08:30-09:25",
    "09:25-10:20",
    "10:30-11:25",
    "11:25-12:20",
    "12:20-13:15",
    "13:15-14:10",
    "14:10-15:05",
    "15:10-16:00",
    "16:00-16:50",
]


def load_time_slots_from_excel(filepath=INPUT_XLSX):
    try:
        sheets = pd.read_excel(filepath, sheet_name=None)
        for sheet_name, sheet in sheets.items():
            if 'time' in sheet_name.lower() or 'slot' in sheet_name.lower():
                for col in sheet.columns:
                    if 'time' in str(col).lower() or 'slot' in str(col).lower():
                        slots = sheet[col].dropna().astype(str).tolist()
                        if slots:
                            return slots
                break
        return _DEFAULT_TIME_SLOTS
    except Exception:
        return _DEFAULT_TIME_SLOTS


TIME_SLOTS = _DEFAULT_TIME_SLOTS
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
DAY_COUNT = len(DAYS)
SLOT_COUNT = len(TIME_SLOTS)
THEORY_SLOTS = list(range(SLOT_COUNT))
LAB_START_SLOTS = [0, 2, 5, 7]


def update_time_slots(filepath=INPUT_XLSX):
    global TIME_SLOTS, SLOT_COUNT, THEORY_SLOTS, LAB_START_SLOTS
    TIME_SLOTS = load_time_slots_from_excel(filepath)
    SLOT_COUNT = len(TIME_SLOTS)
    THEORY_SLOTS = list(range(SLOT_COUNT))
    LAB_START_SLOTS = (
        [0, 2, min(5, SLOT_COUNT - 2), min(7, SLOT_COUNT - 1)] if SLOT_COUNT >= 4 else [0]
    )
    return TIME_SLOTS


# Lunch slot indices — overwritten per request via timingConfig
SY_LUNCH_SLOT = None
TY_LUNCH_SLOT = None

# [Mon..Fri] online-day flags — overwritten per request via onlineConfig
SY_ONLINE_DAYS = [0, 0, 0, 0, 0]
TY_ONLINE_DAYS = [0, 0, 0, 0, 0]

# {"Faculty Name": [[day0_slots], ...]} — 1 = unavailable. Set per request.
FACULTY_AVAILABILITY = {}

# Constraint slot sets — all populated per request from frontend config.
# Empty by default: nothing is locked unless explicitly configured.
TY_MDM_LAB_SLOTS = set()
TY_MDM_THEORY_SLOTS = set()
TY_PEC_LAB_SLOTS = set()
TY_PEC_THEORY_SLOTS = set()
TY_VSEC_LAB_SLOTS = set()
TY_VSEC_THEORY_SLOTS = set()
COMMON_LOCKED_SLOTS = set()


def update_online_config(sy_online=None, ty_online=None):
    global SY_ONLINE_DAYS, TY_ONLINE_DAYS
    if sy_online and isinstance(sy_online, list) and len(sy_online) == 5:
        SY_ONLINE_DAYS = sy_online
    if ty_online and isinstance(ty_online, list) and len(ty_online) == 5:
        TY_ONLINE_DAYS = ty_online


def update_faculty_availability(faculty_avail=None):
    global FACULTY_AVAILABILITY
    if faculty_avail and isinstance(faculty_avail, dict):
        FACULTY_AVAILABILITY = faculty_avail


PEC_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
ONLINE_FILL = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")


PHASE_TIME_LIMIT = 18.0
PHASE_TIME_LIMITS = {
    "PEC": 10.0,
    "VSEC": 30.0,
    "MDM": 15.0,
    "OTHERS": 25.0,
}

try:
    _cpu = os.cpu_count() or 2
except Exception:
    _cpu = 2
NUM_WORKERS = min(8, max(2, _cpu))

# None → different results each run; int → deterministic
SEED = None
