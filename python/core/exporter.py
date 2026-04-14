import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import (
    DAYS,
    PEC_FILL,
    ONLINE_FILL,
    SY_ONLINE_DAYS,
    TY_ONLINE_DAYS,
)


def export_to_excel(assignments, output_xlsx: str) -> bool:
    if not assignments:
        return False
    
    # Re-import online config to get latest values
    import config
    global SY_ONLINE_DAYS, TY_ONLINE_DAYS
    SY_ONLINE_DAYS = config.SY_ONLINE_DAYS
    TY_ONLINE_DAYS = config.TY_ONLINE_DAYS

    df = pd.DataFrame(assignments)
    # Consistent ordering
    df.sort_values(
        by=["DayIndex", "StartSlot", "BatchMain", "Batch", "Course"], inplace=True
    )

    def clean_sheet_name(name, max_len=31):
        """Clean and truncate sheet name for Excel compatibility"""
        # Remove invalid characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        return name[:max_len]

    # Write sheets
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        # Overview sheet
        overview_cols = ["Day", "Time", "Batch", "Course", "Faculty", "Room", "Type"]
        df[overview_cols].to_excel(writer, sheet_name="Overview", index=False)

        # Division-wise sheets (by BatchMain)
        for bmain, g in df.groupby("BatchMain", sort=True):
            cols = ["Day", "Time", "Batch", "Course", "Faculty", "Room", "Type"]
            sheet_name = clean_sheet_name(f"Div_{bmain}")
            g[cols].to_excel(writer, sheet_name=sheet_name, index=False)

        # Faculty-wise sheets
        faculty_list = df["Faculty"].unique()
        for idx, faculty in enumerate(sorted(faculty_list), 1):
            g = df[df["Faculty"] == faculty]
            cols = ["Day", "Time", "Batch", "Course", "Room", "Type"]
            sheet_name = clean_sheet_name(f"Fac_{faculty}")
            g[cols].to_excel(writer, sheet_name=sheet_name, index=False)

        # Room-wise sheets (excluding ONLINE)
        room_df = df[~df["Room"].str.contains("ONLINE", na=False)]
        room_list = room_df["Room"].unique()
        for idx, room in enumerate(sorted(room_list), 1):
            g = room_df[room_df["Room"] == room]
            cols = ["Day", "Time", "Batch", "Course", "Faculty", "Type"]
            sheet_name = clean_sheet_name(f"Room_{room}")
            g[cols].to_excel(writer, sheet_name=sheet_name, index=False)

    # Post-formatting with openpyxl (colors, widths, headers)
    wb = load_workbook(output_xlsx)

    def autofit(ws):
        # Auto column widths
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(10, max_len + 2), 45
            )

        # Header style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

    def paint_rows(ws):
        headers = {
            ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)
        }
        col_course = headers.get("Course")
        col_room = headers.get("Room")
        col_day = headers.get("Day")
        col_batch = headers.get("Batch")
        if not col_course:
            return

        for r in range(2, ws.max_row + 1):
            course_val = ws.cell(row=r, column=col_course).value
            room_val = ws.cell(row=r, column=col_room).value if col_room else None
            day_val = ws.cell(row=r, column=col_day).value if col_day else None
            batch_val = ws.cell(row=r, column=col_batch).value if col_batch else None

            fill = None
            cv = str(course_val).lower() if course_val else ""
            is_pec = ("pec" in cv) or ("professional elective" in cv)
            
            # Check if this session should be online based on day and year
            should_be_online = False
            if day_val and batch_val:
                day_str = str(day_val).strip()
                batch_str = str(batch_val).upper()
                
                # Get day index (0=Monday, 1=Tuesday, etc.)
                day_idx = -1
                for idx, day_name in enumerate(DAYS):
                    if day_name.lower() == day_str.lower():
                        day_idx = idx
                        break
                
                # Check if it's SY or TY and if that day is online
                if day_idx >= 0 and day_idx < 5:  # Only Mon-Fri
                    if "TY" in batch_str or "THIRD" in batch_str:
                        should_be_online = TY_ONLINE_DAYS[day_idx] == 1
                    elif "SY" in batch_str or "SECOND" in batch_str:
                        should_be_online = SY_ONLINE_DAYS[day_idx] == 1
            
            # If PEC is on an online day for TY, it should also be online
            if is_pec:
                fill = PEC_FILL
                if should_be_online:
                    # Override room to ONLINE for PEC on online days
                    if col_room:
                        ws.cell(row=r, column=col_room).value = "ONLINE"
                    room_val = "ONLINE"

            online = (
                room_val and isinstance(room_val, str) and room_val.startswith("ONLINE")
            ) or should_be_online

            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                if fill:
                    cell.fill = fill
                if online and not fill:
                    cell.fill = ONLINE_FILL

    # Format overview
    ws = wb["Overview"]
    autofit(ws)
    paint_rows(ws)

    # Format all other sheets
    for name in wb.sheetnames:
        if name == "Overview":
            continue
        ws = wb[name]
        autofit(ws)
        # Apply coloring to division sheets
        if name.startswith("Div_"):
            paint_rows(ws)

    wb.save(output_xlsx)
    return True
