import re

import openpyxl


REQUIRED_SHEETS = ["Rooms", "Time", "Student", "Faculty", "Subject"]
SLOT_PATTERN = re.compile(r"^\d{2}:\d{2}-\d{2}:\d{2}$")
FACULTY_HEADER = [
    "Name of Faculty",
    "Course Name",
    "Batch",
    "Theory (Hours/week)",
    "Lab(Hours/Week)",
    "Hours",
]


def _result(success, message):
    return {"success": success, "output": message}


def validate_excel_file(filepath):
    """Validate that an Excel workbook matches the scheduler's expected structure.

    Returns a dict {success: bool, output: str}. Never raises — wraps all errors.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return _result(False, f"Could not open workbook: {e}")

    try:
        for name in REQUIRED_SHEETS:
            if name not in wb.sheetnames:
                return _result(
                    False,
                    f"Missing sheet '{name}'. Required: {', '.join(REQUIRED_SHEETS)}",
                )

        # Rooms
        rooms = wb["Rooms"]
        header = [rooms.cell(1, 1).value, rooms.cell(1, 2).value]
        if header != ["Rooms", "Type"]:
            return _result(False, f"Rooms!A1:B1 header must be ['Rooms', 'Type']. Found: {header}")
        for i, row in enumerate(rooms.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] not in ("Theory", "Lab"):
                return _result(False, f"Rooms!B{i} → Type must be 'Theory' or 'Lab'. Found: {row[1]}")

        # Time
        time_sheet = wb["Time"]
        if [time_sheet.cell(1, 1).value, time_sheet.cell(1, 2).value] != ["Index", "Slot"]:
            return _result(False, "Time!A1:B1 header must be ['Index', 'Slot'].")
        for i, row in enumerate(time_sheet.iter_rows(min_row=2, values_only=True), start=2):
            idx, slot = row[0], row[1]
            if not isinstance(idx, int):
                return _result(False, f"Time!A{i} → Index must be an integer. Found: {idx}")
            if not isinstance(slot, str) or not SLOT_PATTERN.match(slot):
                return _result(False, f"Time!B{i} → Slot must look like 08:30-09:25. Found: {slot}")

        # Student
        student = wb["Student"]
        if [student.cell(1, 1).value, student.cell(1, 2).value] != ["Batch", "Type"]:
            return _result(False, "Student!A1:B1 header must be ['Batch', 'Type'].")
        valid_batches = set()
        for i, row in enumerate(student.iter_rows(min_row=2, values_only=True), start=2):
            batch, t = row[0], row[1]
            valid_batches.add(batch)
            if t not in ("Theory", "Lab"):
                return _result(False, f"Student!B{i} → Type must be 'Theory' or 'Lab'. Found: {t}")

        # Subject
        subject = wb["Subject"]
        row1_actual = [subject.cell(1, c).value for c in range(1, 6)]
        if row1_actual != ["Type", "Name", "Hours/Week", "", "Credit"]:
            return _result(
                False,
                f"Subject!Row1 must be ['Type', 'Name', 'Hours/Week', '', 'Credit']. Found: {row1_actual}",
            )
        row2_actual = [subject.cell(2, c).value for c in range(1, 6)]
        if row2_actual != ["", "", "Theory", "Lab", ""]:
            return _result(
                False,
                f"Subject!Row2 must be ['', '', 'Theory', 'Lab', '']. Found: {row2_actual}",
            )
        subject_names = set()
        for i, row in enumerate(subject.iter_rows(min_row=3, values_only=True), start=3):
            _stype, name, th, lab, credit = row[0], row[1], row[2], row[3], row[4]
            subject_names.add(name)
            if th is not None and not isinstance(th, (int, float)):
                return _result(False, f"Subject!C{i} → Theory hours must be numeric. Found: {th}")
            if lab is not None and not isinstance(lab, (int, float)):
                return _result(False, f"Subject!D{i} → Lab hours must be numeric. Found: {lab}")
            if credit is not None and not isinstance(credit, (int, float)):
                return _result(False, f"Subject!E{i} → Credit must be numeric. Found: {credit}")

        # Faculty
        faculty = wb["Faculty"]
        header_row = [faculty.cell(1, c).value for c in range(1, 7)]
        if header_row != FACULTY_HEADER:
            return _result(False, f"Faculty header must be {FACULTY_HEADER}. Found: {header_row}")
        for i, row in enumerate(faculty.iter_rows(min_row=2, values_only=True), start=2):
            _name, course, batch, th, lab, hours = row[0], row[1], row[2], row[3], row[4], row[5]
            if batch not in valid_batches:
                return _result(False, f"Faculty!C{i} → Batch '{batch}' not found in Student sheet.")
            if course not in subject_names:
                return _result(False, f"Faculty!B{i} → Course '{course}' not found in Subject sheet.")
            for col, val in [("D", th), ("E", lab), ("F", hours)]:
                if not isinstance(val, (int, float)):
                    return _result(False, f"Faculty!{col}{i} must be numeric. Found: {val}")

        return _result(True, "Validation passed — file structure is correct.")
    except Exception as e:
        return _result(False, f"Validation error: {e}")
    finally:
        wb.close()
